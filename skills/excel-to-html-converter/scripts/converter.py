import openpyxl
from openpyxl.utils import get_column_letter
import os

def excel_to_html(excel_path, sheet_name, output_path, max_col=None):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found.")
        return
    
    ws = wb[sheet_name]
    
    # If max_col is not provided, try to detect it from the header row
    if max_col is None:
        # Stop at the first empty header if it's beyond column 5 (common data range)
        for c in range(1, ws.max_column + 1):
            header_val = ws.cell(1, c).value
            if c > 5 and header_val is None:
                max_col = c - 1
                break
        if max_col is None:
            max_col = ws.max_column

    # CSS for Professional Look
    # ... (rest of the HTML template remains same)
    html = """
    <!DOCTYPE html>
    <html lang="vi">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Preview: {sheet_name}</title>
        <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600&display=swap" rel="stylesheet">
        <style>
            :root {
                --primary-color: #2563eb;
                --border-color: #e5e7eb;
                --bg-color: #f9fafb;
                --text-color: #1f2937;
            }
            body {
                font-family: 'Inter', sans-serif;
                background-color: var(--bg-color);
                color: var(--text-color);
                margin: 20px;
                line-height: 1.5;
            }
            .container {
                max-width: 100%;
                overflow-x: auto;
                background: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 4px 6px -1px rgb(0 0 0 / 0.1);
            }
            h1 {
                font-size: 1.5rem;
                margin-bottom: 20px;
                color: var(--primary-color);
                border-bottom: 2px solid var(--primary-color);
                padding-bottom: 10px;
            }
            table {
                border-collapse: collapse;
                width: max-content;
                min-width: 100%;
                font-size: 13px;
            }
            th, td {
                border: 1px solid var(--border-color);
                padding: 10px 12px;
                text-align: left;
                white-space: nowrap;
            }
            th {
                background-color: #f3f4f6;
                font-weight: 600;
                position: sticky;
                top: 0;
                z-index: 10;
            }
            tr:hover {
                background-color: #f1f5f9;
            }
            .merged {
                background-color: #fff;
            }
            /* Zebra striping */
            tr:nth-child(even) {
                background-color: #fafafa;
            }
            /* Alignment helper */
            .text-right { text-align: right; }
            .text-center { text-align: center; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>Bản xem trước: {sheet_name}</h1>
            <table>
    """.replace("{sheet_name}", sheet_name)

    merged_cells = ws.merged_cells.ranges
    processed_cells = set()

    for row_idx, row in enumerate(ws.iter_rows(max_col=max_col), 1):
        html += "<tr>"
        for cell in row:
            if cell.coordinate in processed_cells:
                continue
            
            # Check if cell is in merged range
            row_span = 1
            col_span = 1
            is_merged = False
            
            for merged_range in merged_cells:
                if cell.coordinate in merged_range:
                    # Only process the top-left cell of the merged range
                    if cell.row == merged_range.min_row and cell.column == merged_range.min_col:
                        row_span = merged_range.max_row - merged_range.min_row + 1
                        
                        # Clip col_span to max_col
                        actual_max_col = min(merged_range.max_col, max_col)
                        col_span = actual_max_col - merged_range.min_col + 1
                        
                        if col_span <= 0:
                            is_merged = "skip"
                            break

                        is_merged = True
                        
                        # Mark all cells in this range as processed
                        for r in range(merged_range.min_row, merged_range.max_row + 1):
                            for c in range(merged_range.min_col, actual_max_col + 1):
                                processed_cells.add(f"{get_column_letter(c)}{r}")
                    else:
                        is_merged = "skip" # Part of a merged cell but not top-left
                    break
            
            if is_merged == "skip":
                continue
            
            val = cell.value if cell.value is not None else ""
            # Basic sanitization
            if isinstance(val, str):
                val = val.replace("\n", "<br>")
            
            # Formatting (Bold if row is low, etc.)
            style_class = ""
            if isinstance(val, (int, float)):
                style_class = ' class="text-right"'
                if isinstance(val, float):
                    val = f"{val:,.2f}"
            
            tag = "th" if cell.row <= 5 else "td" # Assume first 5 rows are headers
            html += f'<{tag}{style_class} rowspan="{row_span}" colspan="{col_span}">{val}</{tag}>'
            
        html += "</tr>"

    html += """
            </table>
        </div>
    </body>
    </html>
    """
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    try:
        print(f"Successfully converted {sheet_name} to {output_path} (Cols: 1-{max_col})")
    except UnicodeEncodeError:
        print(f"Successfully converted sheet to {output_path} (Cols: 1-{max_col})")

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 4:
        print("Usage: python converter.py <excel_path> <sheet_name> <output_path> [max_col]")
    else:
        m_col = int(sys.argv[4]) if len(sys.argv) > 4 else None
        excel_to_html(sys.argv[1], sys.argv[2], sys.argv[3], max_col=m_col)
