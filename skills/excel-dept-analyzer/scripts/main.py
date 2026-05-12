from openpyxl import load_workbook
import json

FILE = "thống kê chung vượt giờ.xlsx"

wb = load_workbook(FILE, data_only=False)

result = {}


def get_real_cell_value(ws, row, col):

    cell = ws.cell(row, col)

    # cell có value trực tiếp
    if cell.value not in [None, ""]:
        return str(cell.value).strip()

    # check merged cells
    for merged in ws.merged_cells.ranges:

        if (row, col) in merged.cells:

            top_left = ws.cell(
                merged.min_row,
                merged.min_col
            )

            if top_left.value not in [None, ""]:
                return str(top_left.value).strip()

    return None


def get_column_header(ws, row, col):

    headers = []

    # scan nhiều tầng header phía trên
    for r in range(max(1, row - 8), row):

        value = get_real_cell_value(ws, r, col)

        if value:
            headers.append(value)

    # remove duplicate
    headers = list(dict.fromkeys(headers))

    return " > ".join(headers)


def get_row_title(ws, row):

    titles = []

    # scan vài cột đầu
    for c in range(1, 6):

        value = get_real_cell_value(ws, row, c)

        if value:
            titles.append(value)

    titles = list(dict.fromkeys(titles))

    return " | ".join(titles)


for ws in wb.worksheets:

    sheet_result = []

    for row in ws.iter_rows():

        for cell in row:

            value = cell.value

            is_formula = (
                isinstance(value, str)
                and value.startswith("=")
            )

            if not is_formula:
                continue

            item = {
                "sheet": ws.title,
                "cell": cell.coordinate,
                "formula": value,

                "column_header": get_column_header(
                    ws,
                    cell.row,
                    cell.column
                ),

                "row_title": get_row_title(
                    ws,
                    cell.row
                )
            }

            print(item)

            sheet_result.append(item)

    result[ws.title] = sheet_result


with open(
    "formula_with_context_v2.json",
    "w",
    encoding="utf-8"
) as f:

    json.dump(
        result,
        f,
        ensure_ascii=False,
        indent=2
    )

print("DONE")